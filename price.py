from openpyxl import load_workbook

# load excel sheet
wb = load_workbook(filename = 'sure-lok_rip.xlsx')
ws = wb['Sheet1']

# list to hold data
elite = []
premium = []
standard = []

for cell in ws.iter_cols(min_col=4, max_col=4, min_row=2, max_row=485, values_only=True):
  for i in cell: 
    elite.append(i)
print(len(elite))
print("")

for cell in ws.iter_cols(min_col=5, max_col=5, min_row=2, max_row=485, values_only=True):
  for i in cell: 
    premium.append(i)
print(len(premium))
print("")

for cell in ws.iter_cols(min_col=6, max_col=6, min_row=2, max_row=485, values_only=True):
  for i in cell: 
    standard.append(i)
print(len(standard))
print("")

# global row variable to ensure input is in correct row
x = 2 

for y in range(0, 484, 3): 
  print(y)
  # print(elite[y])
  ws.cell(row=x, column=4).value = elite[y]
  ws.cell(row=x, column=5).value = premium[y]
  ws.cell(row=x, column=6).value = standard[y]
  wb.save(filename = 'sure-lok_rip.xlsx')
  x += 1