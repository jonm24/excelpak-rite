from openpyxl import load_workbook

# load excel sheet
wb = load_workbook(filename = 'sure-lok_rip.xlsx')

ws = wb['Sheet1']

# lists to hold data
length = []
width = []
height = []

for cell in ws.iter_cols(min_col=1, max_col=1, min_row=2, max_row=163, values_only=True):
	for i in cell:
		length.append(int(i))	
print(length)
print("")

for cell in ws.iter_cols(min_col=2, max_col=2, min_row=2, max_row=163, values_only=True):
	for i in cell:
		width.append(int(i))	
print(width)
print("")

for cell in ws.iter_cols(min_col=3, max_col=3, min_row=2, max_row=163, values_only=True):
	for i in cell:
		height.append(int(i))	
print(height)
print("")

# global row variable
x = 2

for l in length:
	ws.cell(row=x, column=8).value = "E-" + str(l) + "-" + str(width[x - 2]) + "-" + str(height[x-2])
	ws.cell(row=x + 162, column=8).value = "P-" + str(l) + "-" + str(width[x - 2]) + "-" + str(height[x-2])
	ws.cell(row=x + 324, column=8).value = "S-" + str(l) + "-" + str(width[x - 2]) + "-" + str(height[x-2])
	wb.save(filename = 'sure-lok_rip.xlsx')
	x += 1





