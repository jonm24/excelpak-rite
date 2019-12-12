from openpyxl import load_workbook

# load excel sheet
wb = load_workbook(filename = 'sure-lok_rip.xlsx')

ws = wb['Sheet1']

# lists to hold data
dims = []
yn = []

for cell in ws.iter_cols(min_col=7, max_col=7, min_row=2, max_row=485, values_only=True):
	for i in cell:
		if (i == None):
			continue
		else: 
			dims.append(i)	
print(len(dims))
print("")

for cell in ws.iter_cols(min_col=8, max_col=8, min_row=2, max_row=485, values_only=True):
	for i in cell:
		if (i == None):
			continue
		else: 
			yn.append(i)	
print(len(yn))
print("")

# global row variable
x = 2

for d in dims:
	ws.cell(row=x, column=7).value = d
	ws.cell(row=x, column=8).value = yn[x - 2]
	wb.save(filename = 'sure-lok_rip.xlsx')
	x += 1
