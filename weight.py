from openpyxl import load_workbook

# load excel sheet
wb = load_workbook(filename = 'sure-lok_rip.xlsx')

ws = wb['Sheet1']

# lists to hold data
ps = []
e = []

for cell in ws.iter_cols(min_col=9, max_col=9, min_row=2, max_row=512, values_only=True):
	for i in cell:
		if (i == None):
			continue
		else: 
			ps.append(i)	
print(len(ps))
print("")

for cell in ws.iter_cols(min_col=10, max_col=10, min_row=2, max_row=512, values_only=True):
	for i in cell:
		if (i == None):
			continue
		else: 
			e.append(i)	
print(len(e))
print("")

# global row variable
x = 2

for w in ps:
	ws.cell(row=x, column=9).value = w * 453.592
	ws.cell(row=x, column=10).value = e[x - 2] * 453.592
	wb.save(filename = 'sure-lok_rip.xlsx')
	x += 1

